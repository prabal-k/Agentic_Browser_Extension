"""Format exportable data into various file formats (JSON, CSV, Excel, PDF)."""

import csv
import io
import json
import time


def format_export(
    data: list[dict],
    fmt: str,
    metadata: dict | None = None,
) -> tuple[bytes, str, str]:
    """Format data into the requested file format.

    Args:
        data: List of dicts to export
        fmt: One of 'json', 'csv', 'xlsx', 'pdf'
        metadata: Optional metadata (goal, timestamp, source URL)

    Returns:
        Tuple of (content_bytes, content_type, filename)
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    base_name = f"export_{timestamp}"

    if fmt == "json":
        return _to_json(data, metadata, base_name)
    elif fmt == "csv":
        return _to_csv(data, base_name)
    elif fmt == "xlsx":
        return _to_excel(data, metadata, base_name)
    elif fmt == "pdf":
        return _to_pdf(data, metadata, base_name)
    else:
        # Fallback to JSON
        return _to_json(data, metadata, base_name)


def _to_json(data: list[dict], metadata: dict | None, base_name: str) -> tuple[bytes, str, str]:
    output = {
        "metadata": {
            "exported_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "total_items": len(data),
            **(metadata or {}),
        },
        "items": data,
    }
    content = json.dumps(output, indent=2, ensure_ascii=False).encode("utf-8")
    return content, "application/json", f"{base_name}.json"


def _to_csv(data: list[dict], base_name: str) -> tuple[bytes, str, str]:
    if not data:
        return b"", "text/csv", f"{base_name}.csv"

    # Collect all unique keys across all items
    all_keys: list[str] = []
    seen = set()
    for item in data:
        for key in item:
            if key not in seen:
                all_keys.append(key)
                seen.add(key)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction="ignore")
    writer.writeheader()
    for item in data:
        writer.writerow(item)

    content = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    return content, "text/csv; charset=utf-8", f"{base_name}.csv"


def _to_excel(data: list[dict], metadata: dict | None, base_name: str) -> tuple[bytes, str, str]:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        # Fallback: return CSV if openpyxl not installed
        error = json.dumps({
            "error": "Excel export requires openpyxl. Install with: pip install openpyxl",
            "fallback": "Use CSV or JSON format instead.",
        }).encode("utf-8")
        return error, "application/json", f"{base_name}_error.json"

    if not data:
        return b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{base_name}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    # Collect all keys
    all_keys: list[str] = []
    seen = set()
    for item in data:
        for key in item:
            if key not in seen:
                all_keys.append(key)
                seen.add(key)

    # Header row with styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D5F8A", end_color="2D5F8A", fill_type="solid")

    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, item in enumerate(data, 2):
        for col_idx, key in enumerate(all_keys, 1):
            value = item.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

    # Auto-fit column widths
    for col_idx, key in enumerate(all_keys, 1):
        max_len = len(key)
        for row_idx in range(2, len(data) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, min(len(cell_val), 50))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3

    # Metadata sheet
    if metadata:
        ws_meta = wb.create_sheet("Metadata")
        ws_meta.cell(row=1, column=1, value="Property").font = Font(bold=True)
        ws_meta.cell(row=1, column=2, value="Value").font = Font(bold=True)
        for row_idx, (k, v) in enumerate(metadata.items(), 2):
            ws_meta.cell(row=row_idx, column=1, value=str(k))
            ws_meta.cell(row=row_idx, column=2, value=str(v))

    output = io.BytesIO()
    wb.save(output)
    content = output.getvalue()
    return (
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"{base_name}.xlsx",
    )


def _to_pdf(data: list[dict], metadata: dict | None, base_name: str) -> tuple[bytes, str, str]:
    try:
        from fpdf import FPDF
    except ImportError:
        # Fallback: return CSV if fpdf2 not installed
        error = json.dumps({
            "error": "PDF export requires fpdf2. Install with: pip install fpdf2",
            "fallback": "Use CSV or JSON format instead.",
        }).encode("utf-8")
        return error, "application/json", f"{base_name}_error.json"

    if not data:
        return b"", "application/pdf", f"{base_name}.pdf"

    # Collect keys
    all_keys: list[str] = []
    seen = set()
    for item in data:
        for key in item:
            if key not in seen:
                all_keys.append(key)
                seen.add(key)

    # Limit columns for readability (PDF has limited width)
    display_keys = all_keys[:6]

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 16)
    goal = (metadata or {}).get("goal", "Agent Export")
    if len(goal) > 80:
        goal = goal[:80] + "..."
    pdf.cell(0, 10, goal, ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Exported: {time.strftime('%Y-%m-%d %H:%M:%S')} | Items: {len(data)}", ln=True, align="C")
    pdf.ln(5)

    # Calculate column widths
    page_width = pdf.w - 20  # margins
    col_width = page_width / len(display_keys)

    # Header
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(45, 95, 138)
    pdf.set_text_color(255, 255, 255)
    for key in display_keys:
        label = key.replace("_", " ").title()
        if len(label) > 15:
            label = label[:15]
        pdf.cell(col_width, 8, label, border=1, fill=True, align="C")
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(0, 0, 0)
    for i, item in enumerate(data):
        if i % 2 == 0:
            pdf.set_fill_color(240, 240, 240)
        else:
            pdf.set_fill_color(255, 255, 255)

        for key in display_keys:
            value = str(item.get(key, ""))
            if len(value) > 40:
                value = value[:40] + "..."
            pdf.cell(col_width, 7, value, border=1, fill=True)
        pdf.ln()

    content = bytes(pdf.output())
    return content, "application/pdf", f"{base_name}.pdf"
